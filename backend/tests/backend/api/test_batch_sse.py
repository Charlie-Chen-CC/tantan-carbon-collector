"""
P0-4 回归保护 - batch SSE 真实流式进度

WHY:
  docs/CODE_REVIEW_2026-06-03.md 4.4 + Agent 1 A6 / Agent 4 C3 三 Agent
  交叉验证：extract_router.py extract_batch 端点只 yield 'started' 和
  'complete' 两个事件，中间的 per-file progress 事件从未发出。
  原占位注释：'P0-4 batch SSE 假流式修复见 fix/phase1-p0-4-batch-sse 分支'
  即每个 batch 用户看不到 N 个文件处理进度，体验是「按了按钮卡几秒再
  一次性出结果」，跟本地同步调用无差别。

  修复后，event 序列应为：
    started → progress(1/total) → progress(2/total) → ... → progress(total/total) → complete
  其中 processing 事件数 == total（每个文件一次回调）。

  重要：本测试用 mock 替换 BatchFileProcessor，**不**测文件分组逻辑。
  BatchFileProcessor.process_batch 存在独立 bug（_priority_to_group_key 不
  映射 'excel' 组 → 纯 xlsx 批次 0 处理）属于其他 P 级别，本测试隔离 SSE
  调度本身。P0-4 关注的是「progress_callback 是否真被调用且事件能流出去」。
"""
import ast
import io
import json
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient

from tantan.backend.main import app


def _make_minimal_xlsx(filename: str = "test.xlsx") -> tuple[str, bytes, str]:
    """构造一个最小的合法 xlsx 字节流。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "能源类型"
    ws["A2"] = "电力"
    buf = io.BytesIO()
    wb.save(buf)
    return filename, buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _parse_sse_events(raw: str) -> list[dict]:
    """解析 SSE 原始响应为 [{event, data}, ...] 列表。

    sse_starlette 用 \\r\\n 行尾（RFC 8895 / WHATWG 规范），块分隔是 \\r\\n\\r\\n。
    """
    events = []
    # 统一换行符（防御性，归一为 \n 再处理）
    normalized = raw.replace("\r\n", "\n")
    for block in normalized.split("\n\n"):
        block = block.strip()
        if not block:
            continue
        event_name = None
        data_str = None
        for line in block.split("\n"):
            if line.startswith("event:"):
                event_name = line.split(":", 1)[1].strip()
            elif line.startswith("data:"):
                data_str = line.split(":", 1)[1].strip()
        if event_name and data_str:
            try:
                events.append({"event": event_name, "data": json.loads(data_str)})
            except json.JSONDecodeError:
                events.append({"event": event_name, "data": data_str})
    return events


def _create_session(client: TestClient) -> str:
    resp = client.post("/api/session", json={})
    assert resp.status_code == 200, f"create session failed: {resp.text}"
    return resp.json()["session_id"]


def _mock_process_batch_with_progress():
    """构造一个 mock 替换 BatchFileProcessor.process_batch，按 N 文件触发 progress 事件。"""
    from tantan.backend.agents.file_processor import ProcessingResult

    async def fake_process_batch(self, files, progress_callback=None):
        import asyncio
        for i in range(1, len(files) + 1):
            await asyncio.sleep(0.01)  # 让控制权回到 event loop
            if progress_callback:
                await progress_callback(i, len(files))
        return ProcessingResult(
            status="completed",
            extracted={"电力": {"value": 100, "source": "mock", "priority": 1}},
            warnings=[],
            failed_files=[],
        )
    return fake_process_batch


class TestBatchSSEProgress:
    """P0-4: batch SSE 必须 yield per-file progress 事件"""

    def test_batch_emits_progress_per_file(self, registered_user):
        """上传 5 个文件 → 期望 started + 5 个 processing + complete = 7 个事件"""
        client = registered_user["client"]
        session_id = _create_session(client)

        files_payload = []
        for i in range(5):
            fname, content, ctype = _make_minimal_xlsx(f"file_{i}.xlsx")
            files_payload.append(("files", (fname, content, ctype)))

        with patch(
            "tantan.backend.agents.file_processor.BatchFileProcessor.process_batch",
            _mock_process_batch_with_progress(),
        ):
            with client.stream(
                "POST",
                f"/api/extract/{session_id}/section/1/batch",
                files=files_payload,
            ) as resp:
                assert resp.status_code == 200, f"batch request failed: {resp.status_code}"
                raw = "".join(resp.iter_text())

        print(f"\n[DEBUG TEST] raw SSE:\n{repr(raw)}\n")
        events = _parse_sse_events(raw)
        event_names = [e["event"] for e in events]

        # 必须有：started + 5 个 processing + complete
        assert "complete" in event_names, f"缺少 complete 事件，实际: {event_names}"

        processing_events = [
            e for e in events
            if e["event"] == "progress" and e["data"].get("status") == "processing"
        ]
        assert len(processing_events) == 5, (
            f"应有 5 个 processing 事件（每文件一次），实际 {len(processing_events)} 个: "
            f"{[e['data'] for e in processing_events]}"
        )

        # progress 的 processed/total 必须严格递增到 5/5
        for i, ev in enumerate(processing_events, start=1):
            data = ev["data"]
            assert data["processed"] == i, f"progress #{i} 的 processed 应为 {i}，实际 {data}"
            assert data["total"] == 5, f"progress #{i} 的 total 应为 5，实际 {data}"

    def test_batch_started_event_has_zero_processed(self, registered_user):
        """started 事件是 0/total，与第一个 progress 事件不重复计数"""
        client = registered_user["client"]
        session_id = _create_session(client)

        files_payload = []
        for i in range(3):
            fname, content, ctype = _make_minimal_xlsx(f"start_{i}.xlsx")
            files_payload.append(("files", (fname, content, ctype)))

        with patch(
            "tantan.backend.agents.file_processor.BatchFileProcessor.process_batch",
            _mock_process_batch_with_progress(),
        ):
            with client.stream(
                "POST",
                f"/api/extract/{session_id}/section/1/batch",
                files=files_payload,
            ) as resp:
                raw = "".join(resp.iter_text())

        events = _parse_sse_events(raw)
        started = next(
            e for e in events
            if e["event"] == "progress" and e["data"].get("status") == "started"
        )
        assert started["data"]["processed"] == 0
        assert started["data"]["total"] == 3
        assert started["data"]["status"] == "started"

    def test_batch_complete_event_has_full_result(self, registered_user):
        """complete 事件必须含 status/extracted/warnings/failed_files"""
        client = registered_user["client"]
        session_id = _create_session(client)

        files_payload = []
        for i in range(2):
            fname, content, ctype = _make_minimal_xlsx(f"complete_{i}.xlsx")
            files_payload.append(("files", (fname, content, ctype)))

        with patch(
            "tantan.backend.agents.file_processor.BatchFileProcessor.process_batch",
            _mock_process_batch_with_progress(),
        ):
            with client.stream(
                "POST",
                f"/api/extract/{session_id}/section/1/batch",
                files=files_payload,
            ) as resp:
                raw = "".join(resp.iter_text())

        events = _parse_sse_events(raw)
        complete = next(e for e in events if e["event"] == "complete")
        data = complete["data"]
        assert "status" in data
        assert "extracted" in data
        assert "warnings" in data
        assert "failed_files" in data


class TestBatchSSEASTGuard:
    """P0-4: AST 守门 - 防止假流式占位注释/死代码复发"""

    def test_no_dead_progress_callback_marker(self):
        """extract_router.py 不得再含 'P0-4 batch SSE 假流式修复见' 注释/占位"""
        from pathlib import Path
        path = Path(__file__).resolve().parents[4] / "backend" / "api" / "extract_router.py"
        text = path.read_text(encoding="utf-8")
        assert "P0-4 batch SSE 假流式修复见" not in text, (
            "检测到 P0-4 占位注释残留 — 修复后应已删除此占位说明"
        )
        assert "progress_callback 仍是死代码" not in text

    def test_progress_callback_is_real_callable(self):
        """extract_router.py 的 event_generator 必须把 progress_callback 作为参数传给 process_batch"""
        from pathlib import Path
        path = Path(__file__).resolve().parents[4] / "backend" / "api" / "extract_router.py"
        text = path.read_text(encoding="utf-8")
        assert "progress_callback=" in text, (
            "extract_router.py 未把 progress_callback 传给 process_batch（batch SSE 假流式未修复）"
        )

    def test_sse_yields_encoded_bytes_not_objects(self):
        """所有 yield 必须是 bytes（.encode()），不能 yield ServerSentEvent 对象（StreamingResponse 序列化会抛 TypeError）"""
        from pathlib import Path
        path = Path(__file__).resolve().parents[4] / "backend" / "api" / "extract_router.py"
        text = path.read_text(encoding="utf-8")
        tree = ast.parse(text)
        for node in ast.walk(tree):
            if isinstance(node, ast.AsyncFunctionDef) and node.name == "event_generator":
                for sub in ast.walk(node):
                    if isinstance(sub, ast.Yield) and sub.value is not None:
                        if isinstance(sub.value, ast.Call):
                            called = sub.value.func
                            if isinstance(called, ast.Name) and called.id == "ServerSentEvent":
                                raise AssertionError(
                                    f"event_generator 在 line {sub.lineno} 直接 yield ServerSentEvent 对象，"
                                    f"StreamingResponse 序列化会抛 TypeError（encode() takes 1 positional argument but 2 were given）。"
                                    f"应改为 yield sse.encode() 返回的 bytes。"
                                )
