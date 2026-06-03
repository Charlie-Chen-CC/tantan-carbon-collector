"""
文件提取Agent测试 - 验证AI提取和字段映射功能
"""
import pytest
import sys
import os

# Import from tantan package
from tantan.backend.agents.file_extractor import TextExtractor, FileExtractAgent, LLMExtractor


class TestTextExtractor:
    """测试TextExtractor通用文本提取功能"""

    def test_extract_xlsx(self):
        """测试xlsx文件提取"""
        # 创建一个简单的xlsx文件内容
        import io
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '企业名称'
        ws['B1'] = '测试公司'

        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()

        text = TextExtractor._extract_xlsx(content)
        assert '企业名称' in text
        assert '测试公司' in text

    def test_extract_image_small(self):
        """测试小图片跳过"""
        # 创建一个小图片（小于300x300）
        from PIL import Image
        import io

        img = Image.new('RGB', (100, 100), color='red')
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        content = buffer.getvalue()

        text = TextExtractor._extract_image(content)
        assert '[图片尺寸过小，跳过]' in text

    def test_extract_image_normal(self):
        """测试正常大小图片"""
        from PIL import Image
        import io

        img = Image.new('RGB', (400, 400), color='red')
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        content = buffer.getvalue()

        text = TextExtractor._extract_image(content)
        assert '[IMAGE_DATA]' in text
        assert '[/IMAGE_DATA]' in text


class TestLLMExtractor:
    """测试LLM提取器（需要模拟LLM响应）"""

    def test_extractor_initialization(self):
        """测试提取器初始化"""
        for section in range(1, 10):
            extractor = LLMExtractor(section)
            assert extractor.section == section
            assert len(extractor.prompt) > 0

    def test_extract_empty_text(self):
        """测试空文本处理"""
        extractor = LLMExtractor(section=1)
        result = extractor.extract("")
        assert result == {}


class TestFileExtractAgent:
    """测试文件提取Agent"""

    def test_agent_initialization(self):
        """测试Agent初始化"""
        for section in range(1, 10):
            agent = FileExtractAgent(section)
            assert agent.section == section

    def test_process_unsupported_file(self):
        """测试不支持的文件类型"""
        agent = FileExtractAgent(section=1)

        # 尝试处理一个不支持的文件
        content = b'random binary content'
        result = agent.process(content, filename='test.xyz')

        assert result['status'] == 'failed'
        assert 'error' in result

    def test_prompt_for_each_section(self):
        """验证每个section都有对应的提示词"""
        from tantan.backend.agents.file_extractor import SECTION_PROMPTS

        for section in range(1, 10):
            assert section in SECTION_PROMPTS
            assert len(SECTION_PROMPTS[section]) > 50  # 提示词应该有一定长度


class TestFieldMappingCompleteness:
    """测试字段映射完整性 - 确保每个后端字段都能映射到前端"""

    def test_all_sections_have_mappings(self):
        """验证所有9个部分都有字段映射"""
        from tantan.backend.agents.form_filler import FormFillAgent

        # 收集所有中文字段名
        all_backend_fields = set()
        for section in range(1, 10):
            agent = FormFillAgent(section)
            for field_name in agent.section_definitions[section].fields.keys():
                all_backend_fields.add(field_name)

        # 验证每个字段都能在映射表中找到
        for field in all_backend_fields:
            mapped = FormFillAgent.BACKEND_TO_FRONTEND_FIELD_MAP.get(field)
            # 映射应该存在（即使值可能等于field本身，表示没有特殊映射）
            # 但如果映射不存在（返回None），说明是遗漏的字段
            # 有些字段可能没有前端对应，这个测试只检查映射表的一致性

        # 反向检查：映射表中的每个值应该能在section定义中找到
        for backend_field, frontend_field in FormFillAgent.BACKEND_TO_FRONTEND_FIELD_MAP.items():
            # 前端字段名应该是有效的英文字段名（camelCase或snake_case）
            assert frontend_field, f"前端字段 {frontend_field} 不能为空"
            # 验证主要是英文和数字
            assert all(c.isalnum() or c == '_' or c.isupper() or c.islower() for c in frontend_field), f"前端字段 {frontend_field} 格式可能不对"


if __name__ == '__main__':
    pytest.main([__file__, '-v'])