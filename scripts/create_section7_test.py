#!/usr/bin/env python3
"""
创建 Section 7 (三废处理) 测试数据文件

使用方法:
    python scripts/create_section7_test.py

输出文件:
    test_doc/extractable_by_section/section7/三废处理测试数据.xlsx
"""
import os

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed")
    print("Run: pip install openpyxl")
    exit(1)


def create_section7_test_file():
    """创建包含三废处理数据的测试Excel文件"""

    # 确保输出目录存在
    output_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "test_doc", "extractable_by_section", "section7"
    )
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, "三废处理测试数据.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "三废处理数据"

    # ============================================
    # Section 7 三废处理数据 (对应行号 78-98)
    # ============================================

    # Row 78: 废水处理方式
    ws['A78'] = '废水处理方式'
    ws['B78'] = '厂内有废水处理设施-有厌氧处理工艺单元'

    # Row 79: 废水处理量 (t)
    ws['A79'] = '废水处理量'
    ws['B79'] = 5000

    # Row 80: 目标产品产线废水 (t)
    ws['A80'] = '目标产品产线废水'
    ws['B80'] = 4500

    # Row 81: COD浓度 (mg/L)
    ws['A81'] = 'COD浓度'
    ws['B81'] = 120

    # Row 82: 预留/空行

    # Rows 84-86: 污水处理药剂
    ws['A84'] = '污水处理药剂1'
    ws['B84'] = '聚合氯化铝'

    ws['A85'] = '污水处理药剂2'
    ws['B85'] = '聚丙烯酰胺'

    ws['A86'] = '污水处理药剂3'
    ws['B86'] = ''

    # Row 87: 废气处理方式
    ws['A87'] = '废气处理方式'
    ws['B87'] = '厂内有废气处理设施-RTO焚烧处理'

    # Rows 88-91: 危废处理量
    # 危废委外焚烧
    ws['A88'] = '危废委外焚烧总量'
    ws['B88'] = 15.5
    ws['C88'] = '目标产品产线分解'
    ws['D88'] = 12.0

    # 危废自行焚烧
    ws['A89'] = '危废自行焚烧总量'
    ws['B89'] = 3.2
    ws['C89'] = '目标产品产线分解'
    ws['D89'] = 2.8

    # 危废委外资源化
    ws['A90'] = '危废委外资源化总量'
    ws['B90'] = 8.0
    ws['C90'] = '目标产品产线分解'
    ws['D90'] = 6.5

    # 危废自行资源化
    ws['A91'] = '危废自行资源化总量'
    ws['B91'] = 1.5
    ws['C91'] = '目标产品产线分解'
    ws['D91'] = 1.2

    # Rows 95-98: 烟气处理药剂
    ws['A95'] = '烟气处理药剂1'
    ws['B95'] = '氢氧化钙'

    ws['A96'] = '烟气处理药剂2'
    ws['B96'] = '活性炭'

    ws['A97'] = '烟气处理药剂3'
    ws['B97'] = '尿素'

    ws['A98'] = '烟气处理药剂4'
    ws['B98'] = ''

    # 保存文件
    wb.save(output_path)
    print(f"测试文件已创建: {output_path}")
    return output_path


if __name__ == "__main__":
    create_section7_test_file()